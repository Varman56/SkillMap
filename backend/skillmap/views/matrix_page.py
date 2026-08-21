from urllib.parse import quote

from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from api.helpers import build_skills_cascade_data, build_subcategories_cascade_data
from api.models import Category, Department, User

MAX_LEVEL = 4

# Палитра ячеек экспорта — зеркалит includes/skill_icon.html: фирменный
# фиолетовый #ab4ad5 для подтверждённых уровней (там — сплошная заливка
# иконки независимо от уровня, уровень кодируется формой сектора; в Excel
# формы нет, поэтому уровень кодируем ещё и тоном — светлее для младших
# уровней, ровно #ab4ad5 на уровне 4) и янтарный #D69E2E — для
# неподтверждённого/на рассмотрении (там — контур того же цвета).
_CONFIRMED_FILLS = {
    1: "EAD2F5",
    2: "D5A5EA",
    3: "C077E0",
    4: "AB4AD5",
}
_PENDING_FILL = "FBE6BF"
_HEADER_FILL = "F3F4F6"
_BORDER_COLOR = "D1D5DB"


def _is_hr_or_manager(user) -> bool:
    return user.is_authenticated and user.has_role("HR", "Manager")


def _build_category_columns():
    """Структура колонок: Категория -> Подкатегории -> Навыки."""
    columns = []
    categories = Category.objects.prefetch_related("subcategories__skills").order_by("name")
    for category in categories:
        subcategories_data = []
        for subcategory in sorted(category.subcategories.all(), key=lambda s: s.name):
            skills_data = []
            for skill in sorted(subcategory.skills.all(), key=lambda s: s.name):
                if not skill.is_active:
                    continue
                skills_data.append({"id": skill.id, "name": skill.name})

            if skills_data:
                subcategories_data.append({
                    "name": subcategory.name,
                    "skills": skills_data,
                    "skill_count": len(skills_data) # Количество колонок для colspan подкатегории
                })

        if subcategories_data:
            subcategories_data = sorted(subcategories_data, key=lambda x: x["name"])
            total_category_skills = sum(sub["skill_count"] for sub in subcategories_data)
            columns.append({
                "name": category.name,
                "subcategories": subcategories_data,
                "skill_count": total_category_skills # Количество колонок для colspan категории
            })
    return columns


def _employee_skills_by_id(emp, visible_skill_ids):
    """Сводит до 2 записей UserSkill (подтверждённая + заявка на
    рассмотрении, см. docstring UserSkill) по каждому навыку в одну
    ячейку-словарь для конкретного сотрудника:
      level      — уровень контура иконки (больший из двух, "на что
                   претендует" сотрудник);
      confirmed  — True, если весь навык подтверждён;
      fillLevel  — если задан и меньше level, то до этого уровня
                   подтверждено, а до level — заявка на рассмотрении.
    См. includes/skill_icon.html. Общая функция для matrix_page() и
    matrix_export() — раньше жила только инлайном в matrix_page(), вынесена
    сюда, чтобы у экспорта не завёлся свой отдельный (и потенциально
    расходящийся с экраном) вариант того же самого правила."""
    raw_by_skill = {}
    for us in emp.user_skills.all():
        if us.skill_id in visible_skill_ids:
            raw_by_skill.setdefault(us.skill_id, {})[us.is_approved] = us.level

    skills_by_id = {}
    for skill_id, by_status in raw_by_skill.items():
        approved_level = by_status.get(True)
        pending_level = by_status.get(False)

        if approved_level is not None and pending_level is not None:
            if approved_level >= pending_level:
                # Заявка не выше уже подтверждённого — визуально она
                # целиком "внутри" подтверждённого уровня, отдельно
                # показывать нечего.
                cell = {"level": approved_level, "confirmed": True, "fillLevel": None}
            else:
                cell = {"level": pending_level, "confirmed": False, "fillLevel": approved_level}
        elif approved_level is not None:
            cell = {"level": approved_level, "confirmed": True, "fillLevel": None}
        else:
            cell = {"level": pending_level, "confirmed": False, "fillLevel": None}

        skills_by_id[skill_id] = cell
    return skills_by_id


def _cell_entries(cell):
    """(entries, confirmed_levels) для одной ячейки — зеркалит разбор
    data-level/data-confirmed-levels в applyMatrixFilters() (matrix.html):
    сколько у ячейки "записей" (1 или 2 — см. _employee_skills_by_id) и
    какие из них подтверждены. cell=None — как пустая ячейка на экране
    (level 0, ничего не подтверждено)."""
    if not cell:
        return [0], []
    entries = [cell["level"]]
    confirmed = []
    if cell["confirmed"]:
        confirmed.append(cell["level"])
    fill_level = cell.get("fillLevel")
    if fill_level:
        entries.append(fill_level)
        confirmed.append(fill_level)
    return entries, confirmed


def _cell_visible(entries, confirmed_levels, selected_levels, status_filter):
    """Видна ли ячейка при текущих фильтрах уровня/статуса — зеркалит
    anyMatch в applyMatrixFilters(): достаточно, чтобы ХОТЯ БЫ одна из
    записей ячейки прошла оба условия сразу."""
    for entry in entries:
        if entry not in selected_levels:
            continue
        is_confirmed = entry in confirmed_levels
        if status_filter == "confirmed" and not is_confirmed:
            continue
        if status_filter == "unconfirmed" and is_confirmed:
            continue
        return True
    return False


@login_required(login_url="/login/")
@user_passes_test(_is_hr_or_manager, login_url="/login/")
def matrix_page(request):
    user = request.user

    users_qs = User.objects.select_related("department").prefetch_related(
        "roles",
        "user_skills__skill",
    ).filter(is_active=True, is_intern=False).order_by("full_name")
    # is_active=True — исключает уволенных (is_active=False у уволенных
    # значит "уволен", см. seed_demo_data.py/reserve_page.py). Раньше здесь
    # фильтра не было вообще, поэтому уволенный сотрудник продолжал
    # отображаться в матрице своего бывшего руководителя — несогласованно
    # с "Мой отдел" (department_page.py) и с резервом (только там уволенных
    # можно явно включить обратно галочкой "Только уволенные").
    #
    # is_intern=False — по запросу: практикантов в матрице компетенций
    # видно не должно быть вообще (ни как строку, ни через фильтр — тут
    # его и нет). "isIntern" в employees ниже раньше уходил в JS-данные, но
    # нигде на фронте (matrix.html) не читался — мёртвое поле, оставлено
    # как есть на случай, если пригодится позже.

    if not user.has_role("HR") and user.has_role("Manager"):
        # Без назначенного отдела Manager'у буквально некого показывать
        # (тот же приём, что и в reserve_page.py/ask_page.py).
        users_qs = users_qs.filter(department_id=user.department_id) if user.department_id else users_qs.none()

    all_users = list(users_qs)

    columns = _build_category_columns()
    visible_skill_ids = {s["id"] for c in columns for sub in c["subcategories"] for s in sub["skills"]}

    employees = []
    for emp in all_users:
        skills_by_id = _employee_skills_by_id(emp, visible_skill_ids)

        employees.append(
            {
                "id": emp.id,
                "fullName": emp.full_name,
                "position": emp.position,
                "department": emp.primary_department,
                "role": emp.primary_role,
                "isIntern": emp.is_intern,
                "photo": emp.photo,
                "skills": skills_by_id,
            }
        )

    departments_list = list(Department.objects.values_list("name", flat=True).order_by("name"))
    categories_list = list(Category.objects.values_list("name", flat=True).order_by("name"))

    data = {
        # is_hr — раньше matrix.html решал, показывать ли выпадающий список
        # отделов или залоченное поле, через user.primary_role == "HR" (имя
        # первой по порядку создания роли) — а сам этот view чуть выше решает
        # реальную видимость данных через has_role("HR")/has_role("Manager")
        # (полноценная проверка через M2M, без привязки к порядку). Для
        # пользователя с ОБЕИМИ ролями HR и Manager, у которого Manager
        # технически назначена первой, — бэкенд отдавал данные по всей
        # компании (это HR), а фронт показывал урезанный Manager-интерфейс
        # без выбора отдела: права есть, интерфейса для них нет (аудит,
        # п. 3.2). Передаём готовый is_hr вместо primary_role, вычисленный
        # той же логикой has_role, что и сама фильтрация users_qs выше —
        # шаблон и бэкенд больше не могут разъехаться.
        "is_hr": user.has_role("HR"),
        "departments": departments_list,
        "categories": categories_list,
        # Подкатегории/навыки — списки словарей с data-categories/
        # data-subcategories (через "|") для каскадного сужения фильтров
        # в JS (см. build_subcategories_cascade_data/build_skills_cascade_
        # data в api/helpers.py — общие с reserve_page.py, чтобы дерево
        # "категория -> подкатегория -> навык" считалось одинаково в
        # обоих местах и не расходилось само с собой).
        "subcategories": build_subcategories_cascade_data(),
        "skills": build_skills_cascade_data(),
        "columns": columns,
        "employees": employees,
    }

    return render(request, "matrix.html", data)


@login_required(login_url="/login/")
@user_passes_test(_is_hr_or_manager, login_url="/login/")
def matrix_export(request):
    """GET /matrix/export/ — тот же набор сотрудников/колонок, что и
    matrix_page, но выгруженный в .xlsx.

    Фильтрация матрицы на экране (см. applyMatrixFilters() в matrix.html)
    ПОЛНОСТЬЮ клиентская — сервер о текущем состоянии фильтров ничего не
    знает. Поэтому кнопка "Экспорт" сама собирает GET-параметры из
    текущего состояния фильтров (см. exportMatrix() в extra_js
    matrix.html) и сервер здесь заново применяет ТУ ЖЕ логику видимости
    строк/колонок/ячеек (см. _cell_entries/_cell_visible выше — зеркалят
    anyMatch в applyMatrixFilters()), чтобы файл соответствовал тому, что
    человек видит на экране в момент клика.

    Ячейки, не прошедшие фильтр уровня/статуса, остаются в файле пустыми
    (на экране — просто притушены, opacity/grayscale; в Excel аналога
    "притушить" нет, поэтому решили просто не печатать значение).
    """
    user = request.user

    users_qs = User.objects.select_related("department").prefetch_related(
        "user_skills__skill",
    ).filter(is_active=True, is_intern=False).order_by("full_name")
    # is_intern=False — та же причина, что и в matrix_page() выше: экспорт
    # должен показывать ровно то, что видно на экране, а практикантов там
    # больше нет.

    if not user.has_role("HR") and user.has_role("Manager"):
        users_qs = users_qs.filter(department_id=user.department_id) if user.department_id else users_qs.none()
    elif user.has_role("HR"):
        # Фильтр по отделу — только для HR (Manager и так видит только свой
        # отдел, см. выше и matrix_page()); параметр от Manager'а, даже
        # если бы пришёл, здесь просто не учитывается.
        department_param = (request.GET.get("department") or "").strip()
        if department_param:
            users_qs = users_qs.filter(department__name=department_param)

    all_users = list(users_qs)

    search = (request.GET.get("search") or "").strip().lower()
    if search:
        all_users = [
            u for u in all_users
            if search in u.full_name.lower() or search in (u.position or "").lower()
        ]

    columns = _build_category_columns()
    visible_skill_ids = {s["id"] for c in columns for sub in c["subcategories"] for s in sub["skills"]}

    # Сужаем НАБОР ОТОБРАЖАЕМЫХ колонок под фильтры категории/подкатегории/
    # навыка — ТЕ ЖЕ условия, что и в JS (applyMatrixFilters() прячет целые
    # столбцы, если их категория/подкатегория/навык не совпадает с
    # выбором). visible_skill_ids при этом НЕ сужаем — он определяет, какие
    # UserSkill вообще попадают в ячейку сотрудника (как и на экране, это
    # не зависит от того, какие колонки сейчас видны).
    category_param = (request.GET.get("category") or "").strip()
    subcategory_param = (request.GET.get("subcategory") or "").strip()
    skill_param = (request.GET.get("skill") or "").strip()

    skill_columns = []  # плоский список (category_name, subcategory_name, skill_dict)
    for cat in columns:
        if category_param and cat["name"] != category_param:
            continue
        for sub in cat["subcategories"]:
            if subcategory_param and sub["name"] != subcategory_param:
                continue
            for skill in sub["skills"]:
                if skill_param and skill["name"] != skill_param:
                    continue
                skill_columns.append((cat["name"], sub["name"], skill))

    status_filter = (request.GET.get("status") or "all").strip()
    if status_filter not in ("all", "confirmed", "unconfirmed"):
        status_filter = "all"

    # levels_param может быть пустой строкой "" — это значит "снят каждый
    # чекбокс уровня" (см. matrix.html: params.set('levels', ...join(','))
    # шлётся, если выбрано не все 5), и в таком случае экспорт должен
    # совпасть с картинкой на экране (ничего не подсвечено) — то есть
    # selected_levels должен остаться ПУСТЫМ множеством, а не "все уровни".
    # Старая проверка `if levels_param:` считала "" ложным значением и
    # ошибочно трактовала явно пустой выбор так же, как отсутствие
    # параметра вообще (когда на экран открыли без единого фильтра) —
    # аудит, п. 2.2. None (параметра нет в query string совсем) — это
    # единственный случай, где действительно нужны все уровни по умолчанию.
    levels_param = request.GET.get("levels")
    if levels_param is not None:
        selected_levels = {int(v) for v in levels_param.split(",") if v.strip().isdigit()}
    else:
        selected_levels = {0, 1, 2, 3, 4}

    workbook = _build_matrix_workbook(all_users, skill_columns, visible_skill_ids, status_filter, selected_levels)

    filename = f"Матрица компетенций {timezone.now().strftime('%Y-%m-%d')}.xlsx"
    quoted_filename = quote(filename)
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=\"{quoted_filename}\"; filename*=UTF-8''{quoted_filename}"
    )
    workbook.save(response)
    return response


def _build_matrix_workbook(users, skill_columns, visible_skill_ids, status_filter, selected_levels):
    """Строит .xlsx: 3 строки шапки (категория -> подкатегория -> навык,
    объединённые ячейки — структура похожа на референс-скриншот "базфактор",
    но со строками = реальные сотрудники, а не абстрактные роли), затем по
    строке на сотрудника, дальше — легенда цветов. Никаких формул — это
    статический слепок данных на момент экспорта, поэтому пересчёт
    (recalc.py) не требуется."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Матрица компетенций"

    header_font = Font(name="Arial", bold=True, size=10)
    label_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color=_BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    rotated = Alignment(horizontal="center", vertical="bottom", text_rotation=90, wrap_text=False)
    left_align = Alignment(horizontal="left", vertical="center")

    lead_cols = ["Сотрудник", "Должность", "Отдел"]
    lead_col_count = len(lead_cols)
    first_skill_col = lead_col_count + 1

    for i, label in enumerate(lead_cols):
        col = i + 1
        ws.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
        c = ws.cell(row=1, column=col, value=label)
        c.font = header_font
        c.alignment = center
        for r in (1, 2, 3):
            xc = ws.cell(row=r, column=col)
            xc.fill = header_fill
            xc.border = border

    def _header_groups(key_fn):
        """Группирует skill_columns по подряд идущим одинаковым
        key_fn(...) — та же логика группировки, что и в
        _merge_header_row ниже, но вынесена отдельно, чтобы посчитать
        нужную ширину колонок ДО того, как они будут отрисованы
        (см. _min_span_width и col_widths ниже)."""
        groups = []
        i = 0
        while i < len(skill_columns):
            key = key_fn(skill_columns[i])
            j = i
            while j < len(skill_columns) and key_fn(skill_columns[j]) == key:
                j += 1
            groups.append((i, j - i, key))
            i = j
        return groups

    def _min_span_width(label, col_count):
        """Минимальная суммарная ширина (в юнитах Excel column width) для
        объединённой ячейки шапки категории/подкатегории. Раньше все
        навык-колонки были фиксированной шириной 4.5 — этого хватает для
        коротких названий, но длинное однословное название (например
        "Контейнеризация", "Оркестрация" — без пробелов, переносить
        по словам некуда) при узкой колонке резалось Excel на 3-4
        нечитаемых обрывка по 3 символа. Оцениваем грубо (эмпирический
        коэффициент под Arial bold 10pt): короткие названия (<=10
        символов) — в одну строку, длинные — переносим максимум на 2
        строки. Точность до пикселя не нужна — единственная цель:
        избежать нечитаемых мелких обрывков."""
        length = len(label)
        chars_per_line = length if length <= 10 else -(-length // 2)
        needed_total = chars_per_line * 1.3
        return max(4.5, needed_total / col_count)

    col_widths = [4.5] * len(skill_columns)
    for start, span, label in _header_groups(lambda t: t[0]):
        w = _min_span_width(label, span)
        for k in range(span):
            col_widths[start + k] = max(col_widths[start + k], w)
    for start, span, label in _header_groups(lambda t: t[1]):
        w = _min_span_width(label, span)
        for k in range(span):
            col_widths[start + k] = max(col_widths[start + k], w)
    # Ширина колонки раньше считалась ТОЛЬКО по названию категории/
    # подкатегории (merge-заголовки строк 1-2) — само название навыка
    # (строка 3, повёрнутое на 90°) в расчёт не входило вообще. Для
    # категории/подкатегории с широким merge (много навыков в span) это
    # было не видно, но у навыка с длинным «безпробельным» названием
    # (например "Контейнеризация") при узком span'е колонка всё равно
    # оставалась минимальной ширины и текст обрезался тем же образом,
    # что чинили выше для категорий (аудит, п. 4.4). span=1, потому что
    # колонка навыка никогда не объединяется с соседними (в отличие от
    # категории/подкатегии) — та же _min_span_width, что и для них.
    for offset, (_, _, skill) in enumerate(skill_columns):
        w = _min_span_width(skill["name"], 1)
        col_widths[offset] = max(col_widths[offset], w)

    def _merge_header_row(row_idx, key_fn):
        idx = first_skill_col
        i = 0
        while i < len(skill_columns):
            key = key_fn(skill_columns[i])
            j = i
            while j < len(skill_columns) and key_fn(skill_columns[j]) == key:
                j += 1
            span = j - i
            if span > 1:
                ws.merge_cells(start_row=row_idx, start_column=idx, end_row=row_idx, end_column=idx + span - 1)
            c = ws.cell(row=row_idx, column=idx, value=key)
            c.font = header_font
            c.alignment = center
            for k in range(span):
                xc = ws.cell(row=row_idx, column=idx + k)
                xc.fill = header_fill
                xc.border = border
            idx += span
            i = j

    # Строка 1 — категории, строка 2 — подкатегории (обе объединяются по
    # соседним колонкам с одинаковым именем — колонки уже сгруппированы
    # именно так, см. _build_category_columns()).
    _merge_header_row(1, lambda t: t[0])
    _merge_header_row(2, lambda t: t[1])

    # Строка 3 — сами навыки, текст повёрнут на 90° и колонки узкие (как на
    # референс-скриншоте "базфактор") — иначе десятки навыков в ширину
    # были бы нечитаемы.
    for offset, (_, _, skill) in enumerate(skill_columns):
        col = first_skill_col + offset
        c = ws.cell(row=3, column=col, value=skill["name"])
        c.font = header_font
        c.alignment = rotated
        c.fill = header_fill
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = col_widths[offset]

    # Категории/подкатегории с узким набором навыков дают узкий merge (см.
    # column width=4.5 у самих навыков ниже) — при коротком названии
    # текст переносится на 2 строки, поэтому строкам шапки нужна высота
    # чуть больше однострочной, иначе вторая строка обрежется по границе
    # ячейки (было 22 — мало для, например, "Soft Skills" в 2 строки).
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 34
    ws.row_dimensions[3].height = 120

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18

    row = 4
    for emp in users:
        skills_by_id = _employee_skills_by_id(emp, visible_skill_ids)

        for col, value in ((1, emp.full_name), (2, emp.position or "—"), (3, emp.primary_department or "—")):
            c = ws.cell(row=row, column=col, value=value)
            c.font = label_font
            c.alignment = left_align
            c.border = border

        for offset, (_, _, skill) in enumerate(skill_columns):
            col = first_skill_col + offset
            cell_data = skills_by_id.get(skill["id"])
            entries, confirmed_levels = _cell_entries(cell_data)
            visible = _cell_visible(entries, confirmed_levels, selected_levels, status_filter)

            xc = ws.cell(row=row, column=col)
            xc.border = border
            xc.alignment = center

            if visible and cell_data:
                level = cell_data["level"]
                fill_level = cell_data.get("fillLevel")
                if fill_level:
                    # Смешанная ячейка — подтверждён fill_level, заявка на
                    # рассмотрении до level (см. _employee_skills_by_id).
                    # Диагональная штриховка двумя цветами сразу
                    # (подтверждённый оттенок фиолетового + янтарный).
                    # Было "darkUp" (фиолетовый — основной, янтарные линии
                    # — тонкие) — на уровне 3-4 фиолетовый почти сплошной,
                    # полосы не видны и тёмный текст на тёмном фоне не
                    # читался. Взяли "lightUp" — наоборот, основной фон
                    # светлый янтарный, а фиолетовые линии тонкие поверх
                    # него: полосы видно на любом уровне, а фон всегда
                    # светлый — тёмный текст остаётся читаемым без доп.
                    # условий на уровень.
                    xc.value = f"{fill_level}→{level}"
                    xc.fill = PatternFill(
                        patternType="lightUp",
                        fgColor=_CONFIRMED_FILLS.get(fill_level, _CONFIRMED_FILLS[4]),
                        bgColor=_PENDING_FILL,
                    )
                    xc.font = Font(name="Arial", size=9, bold=True, color="1F2937")
                elif cell_data["confirmed"]:
                    xc.value = level
                    xc.fill = PatternFill("solid", fgColor=_CONFIRMED_FILLS.get(level, _CONFIRMED_FILLS[4]))
                    xc.font = Font(name="Arial", size=9, bold=True, color="FFFFFF" if level >= 3 else "3B0764")
                else:
                    xc.value = level
                    xc.fill = PatternFill("solid", fgColor=_PENDING_FILL)
                    xc.font = Font(name="Arial", size=9, bold=True, color="7C4A03")
            # иначе ячейка остаётся пустой: либо уровень 0 (навыка нет —
            # на экране это просто пустой контур, без текста), либо запись
            # есть, но не проходит текущий фильтр уровня/статуса (на
            # экране — притушена, см. docstring выше).
        row += 1

    # Легенда — под таблицей, чтобы открывший файл понимал цвета без
    # обращения к самому приложению.
    legend_row = row + 2
    lc = ws.cell(row=legend_row, column=1, value="Легенда")
    lc.font = Font(name="Arial", bold=True, size=11)
    legend_row += 1

    legend_items = [
        ("Подтверждён — уровень 1", PatternFill("solid", fgColor=_CONFIRMED_FILLS[1])),
        ("Подтверждён — уровень 2", PatternFill("solid", fgColor=_CONFIRMED_FILLS[2])),
        ("Подтверждён — уровень 3", PatternFill("solid", fgColor=_CONFIRMED_FILLS[3])),
        ("Подтверждён — уровень 4", PatternFill("solid", fgColor=_CONFIRMED_FILLS[4])),
        ("На рассмотрении / не подтверждён", PatternFill("solid", fgColor=_PENDING_FILL)),
        (
            # Штриховка тем же цветом, что и в таблице — тут показываем
            # уровень 4 просто для примера, у конкретной ячейки цвет полос
            # будет соответствовать её собственному подтверждённому уровню.
            "Двойной статус — заливка «в полоску» (см. «X→Y» ниже)",
            PatternFill(patternType="lightUp", fgColor=_CONFIRMED_FILLS[4], bgColor=_PENDING_FILL),
        ),
    ]
    for label, fill in legend_items:
        swatch = ws.cell(row=legend_row, column=1)
        swatch.fill = fill
        swatch.border = border
        text_cell = ws.cell(row=legend_row, column=2, value=label)
        text_cell.font = label_font
        text_cell.alignment = left_align
        legend_row += 1

    note = ws.cell(
        row=legend_row,
        column=1,
        value="«X→Y» — подтверждён уровень X (цвет полос), заявка на уровень Y ещё на рассмотрении.",
    )
    note.font = Font(name="Arial", italic=True, size=9)
    legend_row += 1

    if not skill_columns:
        ws.cell(row=legend_row, column=1, value="По выбранным фильтрам не найдено ни одного навыка.").font = label_font
        legend_row += 1
    if not users:
        ws.cell(row=legend_row, column=1, value="По выбранным фильтрам не найдено ни одного сотрудника.").font = label_font

    ws.freeze_panes = ws.cell(row=4, column=first_skill_col).coordinate

    # Печать — не основной сценарий (файл в первую очередь открывают в
    # Excel и скроллят), но раз уж навыков может быть много колонок,
    # пусть хотя бы вписывается по ширине листа, если кто-то распечатает.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    return wb